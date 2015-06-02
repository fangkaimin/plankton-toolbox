#!/usr/bin/env python
# -*- coding:utf-8 -*-
#
# Copyright (c) 2010-2015 SMHI, Swedish Meteorological and Hydrological Institute 
# License: MIT License (see LICENSE.txt or http://opensource.org/licenses/mit).
#
from __future__ import unicode_literals

import envmonlib
try: 
    openpyxl_installed = False
    import openpyxl
    openpyxl_installed = True
except ImportError:
    print('Python package openpyxl is missing. Please install.')

class ExcelFiles():
    """ """
    def __init__(self):
        """ """
        self._metadata = {}
        self._header = [] # 
        self._rows = []
        
    def readToTableDataset(self, 
                           file_name,
                           sheet_name = None, 
                           header_row = 0, 
                           data_rows_from = 1, 
                           data_rows_to = None): # None = read all.
        """ """
        target_header = [] 
        target_rows = []
        #
        if openpyxl_installed == False:
            raise UserWarning('The python package 'openpyxl' is not installed. Can\'t read .xlsx files.')
        if file_name == None:
            raise UserWarning('File name is missing.')
        try:
            workbook = openpyxl.load_workbook(file_name, use_iterators = True) # Supports big files.
            if workbook == None:
                raise UserWarning('Can\'t read Excel (.xlsx) file.')
            worksheet = None
            if sheet_name:
                # 
                if sheet_name in workbook.get_sheet_names():
                    worksheet = workbook.get_sheet_by_name(name = sheet_name)
                else:
                    raise UserWarning('Excel sheet ' + sheet_name + " not available.')      
            else:
                # Use the first sheet if not specified.
                worksheet = workbook.get_sheet_by_name(name = workbook.get_sheet_names()[0])
            #
            header = []
            for rowindex, row in enumerate(worksheet.iter_rows()): ### BIG.
                if data_rows_to and (rowindex > data_rows_to):
                    break # Break loop if data_row_to is defined and exceeded.
                elif rowindex == header_row:
                    for cell in row:
###                         value = cell.internal_value
                        value = cell.value
                        if value == None:
                            header.append('')
                        else:
                            header.append(unicode(value).strip())
                    #
                    target_header = header
                elif rowindex >= data_rows_from:
                    newrow = []
                    for cell in row: ### BIG.
###                        value = cell.internal_value
                        value = cell.value
                        if value == None:
                            newrow.append('')
                        else:
                            newrow.append(unicode(value).strip())
                    target_rows.append(newrow)
            #
            return (target_header, target_rows)

        #  
        except Exception as e:
            print('Can\'t read Excel file. File name: ' + file_name )
            raise UserWarning('Can\'t read Excel file. File name: ' + file_name + '. Exception: ' + unicode(e))

#    def readToTableDataset(self, 
#                           target_dataset, 
#                           file_name,
#                           sheet_name = None, 
#                           header_row = 0, 
#                           data_rows_from = 1, 
#                           data_rows_to = None, # None = read all.
#                           used_columns_from = 0, 
#                           used_columns_to = None): # None = read all.
#        """ """
#        if file_name == None:
#            raise UserWarning('File name is missing.')
#        if not isinstance(target_dataset, envmonlib.DatasetTable):
#            raise UserWarning('Target dataset is not of valid type.')
#        try:
#            workbook = excelreader.load_workbook(file_name)
#            if workbook == None:
#                raise UserWarning('Can\'t read Excel (.xlsx) file.')
#            worksheet = None
#            if sheet_name:
#                # 
#                if sheet_name in workbook.get_sheet_names():
#                    worksheet = workbook.get_sheet_by_name(sheet_name)
#                else:
#                    raise UserWarning('Excel sheet ' + sheet_name + " not available.')      
#            else:
#                # Use the first sheet if not specified.
#                worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
#            # Prepare boundary. 
#            if used_columns_to:
#                used_columns_to = min(used_columns_to,  worksheet.get_highest_column())
#            else:
#                used_columns_to = worksheet.get_highest_column()
#            if data_rows_to:
#                data_rows_to = min(data_rows_to,  worksheet.get_highest_row())
#            else:
#                data_rows_to = worksheet.get_highest_row()
#            # Read header row.
#            header = []
#            for col in xrange(used_columns_from, worksheet.get_highest_column()):
#                value = worksheet.cell(row=header_row, column=col).value
#                if value:
#                    header.append(unicode(value).strip())
#                else:
#                    header.append('')
#            target_dataset.setHeader(header)
#            # Read data rows.
#            for row in xrange(data_rows_from, data_rows_to): 
#                newrow = []
#                for col in xrange(used_columns_from, worksheet.get_highest_column()):
#                    value = worksheet.cell(row=row, column=col).value
#                    if value:
#                        newrow.append(unicode(value).strip())
#                    else:
#                        newrow.append('')
#                target_dataset.appendRow(newrow)
#        #  
#        except Exception:
#            print('Can\'t read Excel file. File name: ' + file_name )
#            raise UserWarning('Can\'t read Excel file. File name: ' + file_name)

    def writeTableDataset(self, table_dataset, file_name):
        """ """
        if file_name == None:
            raise UserWarning('File name is missing.')
        if not isinstance(table_dataset, envmonlib.DatasetTable):
            raise UserWarning('Dataset is not of a valid type.')
        try:
            workbook =  openpyxl.Workbook(optimized_write = True)  # Supports big files.
            worksheet = workbook.create_sheet()
            # Header.
            worksheet.append(table_dataset.getHeader())
            # Rows.
            for row in table_dataset.getRows():
                worksheet.append(row)
            # Save to file.   
            workbook.save(file_name)
        #
        except (IOError, OSError):
            envmonlib.Logging().log('Failed to write to file: ' + file_name)
            raise UserWarning('Failed to write to file: ' + file_name)

#    def writeTableDataset(self, table_dataset, file_name):
#        """ """
#        if file_name == None:
#            raise UserWarning('File name is missing.')
#        if not isinstance(table_dataset, envmonlib.DatasetTable):
#            raise UserWarning('Dataset is not of a valid type.')
#        try:
#            workbook = excelworkbook.Workbook() # Create workbook.
#            worksheet = workbook.get_active_sheet() # Workbooks contains at least one worksheet.
#            # Header.
#            for columnindex, item in enumerate(table_dataset.getHeader()):
#                cell = worksheet.cell(row=0, column=columnindex)
#                cell.value = unicode(item)
#            # Rows.
#            for rowindex, row in enumerate(table_dataset.getRows()):
#                for columnindex, item in enumerate(row):
#                    cell = worksheet.cell(row=rowindex + 1, column=columnindex)
#                    cell.value = unicode(item)
#            #    
#            excelwriter.save_workbook(workbook, file_name)
#            # Close.
#            # TODO: Not needed?        
#        except (IOError, OSError):
#            envmonlib.Logging().log('Failed to write to file: ' + file_name)
#            raise UserWarning('Failed to write to file: ' + file_name)
